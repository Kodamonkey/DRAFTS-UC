#!/usr/bin/env python3
"""
Script para analizar validación completa de fases ALMA.

Compara los 8 pulsos canónicos + 54 extras del Excel con las detecciones
automáticas de cada fase metodológica, calcula métricas de rendimiento
y genera archivos Excel ordenados por relevancia.
"""

import re
import pandas as pd
import numpy as np
import yaml
from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Importar funciones del script anterior
import sys
sys.path.append(str(Path(__file__).parent))
from compare_validated_with_detections import (
    normalize_filename,
    filter_special_files
)


def load_config(config_path: str) -> Dict:
    """Carga la configuración desde config.yaml"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def load_validated_candidates(excel_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Carga y separa los candidatos validados en canónicos (8) y extras (54).
    
    Args:
        excel_path: Ruta al archivo tabla_transformada.xlsx
        
    Returns:
        Tuple de (canonicos_df, extras_df)
    """
    print(f"Cargando candidatos validados desde: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name='Datos')
    
    # Limpiar filas vacías o con encabezados duplicados
    df = df.dropna(subset=['Folder', 'nombre_archivo'], how='all')
    df = df[df['Folder'] != 'Folder']  # Remover encabezados duplicados
    
    # Normalizar nombres de archivo
    if 'nombre_archivo' in df.columns:
        df['nombre_archivo_normalized'] = df['nombre_archivo'].apply(normalize_filename)
    else:
        print("⚠ Advertencia: No se encontró la columna 'nombre_archivo'")
        df['nombre_archivo_normalized'] = ''
    
    # Normalizar campo 'candidato tiempo' - remover texto entre paréntesis
    if 'candidato tiempo' in df.columns:
        df['candidato_tiempo_clean'] = df['candidato tiempo'].astype(str).apply(
            lambda x: re.sub(r'\s*\([^)]*\)', '', x).strip()
        )
        df['candidato_tiempo_clean'] = pd.to_numeric(
            df['candidato_tiempo_clean'], 
            errors='coerce'
        )
    else:
        print("⚠ Advertencia: No se encontró la columna 'candidato tiempo'")
        df['candidato_tiempo_clean'] = None
    
    # Identificar canónicos: buscar "(pulso jose)" en candidato tiempo
    df['is_canonical'] = df['candidato tiempo'].astype(str).str.contains(
        'pulso jose', case=False, na=False
    )
    
    # Separar canónicos y extras
    canonicos_df = df[df['is_canonical']].copy()
    extras_df = df[~df['is_canonical']].copy()
    
    # Limpiar filas vacías de extras
    extras_df = extras_df.dropna(subset=['nombre_archivo', 'candidato tiempo'], how='all')
    
    if len(canonicos_df) != 8:
        print(f"  [WARNING] Se encontraron {len(canonicos_df)} canónicos (esperados 8)")
    
    print(f"  [OK] Cargados {len(canonicos_df)} canónicos y {len(extras_df)} extras")
    return canonicos_df, extras_df


def load_detected_candidates(csv_path: str) -> pd.DataFrame:
    """
    Carga y normaliza los candidatos detectados automáticamente.
    
    Args:
        csv_path: Ruta al archivo combined_candidates-*.csv
        
    Returns:
        DataFrame con candidatos detectados normalizados
    """
    print(f"Cargando candidatos detectados desde: {csv_path}")
    df = pd.read_csv(csv_path, encoding='utf-8')
    
    # Filtrar archivos especiales con tiempos en el nombre
    df = filter_special_files(df)
    
    # Normalizar campo 'file' para matching consistente
    if 'file' in df.columns:
        df['file_clean'] = df['file'].apply(normalize_filename)
    else:
        print("⚠ Advertencia: No se encontró la columna 'file'")
        df['file_clean'] = ''
    
    # Asegurar que t_sec_dm_time sea numérico
    if 't_sec_dm_time' in df.columns:
        df['t_sec_dm_time'] = pd.to_numeric(df['t_sec_dm_time'], errors='coerce')
    else:
        print("⚠ Advertencia: No se encontró la columna 't_sec_dm_time'")
        df['t_sec_dm_time'] = None
    
    print(f"  [OK] Cargados {len(df)} candidatos detectados (después de filtrar archivos especiales)")
    return df


def find_matches(validated_df: pd.DataFrame, detected_df: pd.DataFrame, 
                 source_name: str) -> pd.DataFrame:
    """
    Encuentra matches entre candidatos validados y detectados.
    También incluye candidatos detectados que no tienen match (etiquetados como "new").
    
    Args:
        validated_df: DataFrame con candidatos validados (canónicos o extras)
        detected_df: DataFrame con candidatos detectados
        source_name: Nombre del archivo fuente
        
    Returns:
        DataFrame con todos los matches y candidatos nuevos
    """
    results = []
    matched_detected_indices = set()
    
    # Procesar todos los candidatos validados
    for idx, validated_row in validated_df.iterrows():
        nombre_archivo_normalized = validated_row.get('nombre_archivo_normalized', '')
        candidato_tiempo = validated_row.get('candidato_tiempo_clean')
        
        if pd.isna(nombre_archivo_normalized) or nombre_archivo_normalized == '' or pd.isna(candidato_tiempo):
            result_row = validated_row.to_dict()
            result_row['has_match'] = False
            result_row['num_matches'] = 0
            result_row['source_file'] = source_name
            result_row['match_type'] = 'validated'
            for col in detected_df.columns:
                if col not in ['file_clean']:
                    result_row[f'detected_{col}'] = None
            results.append(result_row)
            continue
        
        # Filtrar por archivo usando nombres normalizados
        file_matches = detected_df[
            detected_df['file_clean'] == nombre_archivo_normalized
        ].copy()
        
        if len(file_matches) == 0:
            result_row = validated_row.to_dict()
            result_row['has_match'] = False
            result_row['num_matches'] = 0
            result_row['source_file'] = source_name
            result_row['match_type'] = 'validated'
            for col in detected_df.columns:
                if col not in ['file_clean']:
                    result_row[f'detected_{col}'] = None
            results.append(result_row)
            continue
        
        # Filtrar por tiempo (diferencia < 0.1 segundos)
        file_matches['time_diff'] = abs(file_matches['t_sec_dm_time'] - candidato_tiempo)
        time_matches = file_matches[file_matches['time_diff'] < 0.1].copy()
        
        num_matches = len(time_matches)
        
        if num_matches == 0:
            result_row = validated_row.to_dict()
            result_row['has_match'] = False
            result_row['num_matches'] = 0
            result_row['source_file'] = source_name
            result_row['match_type'] = 'validated'
            for col in detected_df.columns:
                if col not in ['file_clean', 'time_diff']:
                    result_row[f'detected_{col}'] = None
            results.append(result_row)
        else:
            # Hay matches - crear una fila por cada match
            for match_idx, match_row in time_matches.iterrows():
                matched_detected_indices.add(match_idx)
                result_row = validated_row.to_dict()
                result_row['has_match'] = True
                result_row['num_matches'] = num_matches
                result_row['source_file'] = source_name
                result_row['match_type'] = 'validated'
                for col in detected_df.columns:
                    if col not in ['file_clean', 'time_diff']:
                        result_row[f'detected_{col}'] = match_row[col]
                results.append(result_row)
    
    # Agregar candidatos detectados que NO tienen match (etiquetados como "new")
    unmatched_detected = detected_df[~detected_df.index.isin(matched_detected_indices)].copy()
    
    for idx, detected_row in unmatched_detected.iterrows():
        result_row = {}
        # Columnas validated_ vacías
        for col in validated_df.columns:
            if col not in ['candidato_tiempo_clean', 'nombre_archivo_normalized', 'is_canonical']:
                if col.startswith('validated_'):
                    result_row[col] = None
                elif col == 'Folder':
                    result_row['validated_Folder'] = None
                elif col == 'subfolder':
                    result_row['validated_subfolder'] = None
                elif col == 'nombre_archivo':
                    result_row['validated_nombre_archivo'] = None
                elif col == 'candidato tiempo':
                    result_row['validated_candidato_tiempo'] = None
                elif col == 'SNR':
                    result_row['validated_SNR'] = None
                elif col == 'choosen':
                    result_row['validated_choosen'] = None
                else:
                    result_row[f'validated_{col}'] = None
        
        result_row['has_match'] = False
        result_row['num_matches'] = 0
        result_row['source_file'] = source_name
        result_row['match_type'] = 'new'
        
        # Agregar todas las columnas detected_
        for col in detected_df.columns:
            if col not in ['file_clean', 'time_diff']:
                result_row[f'detected_{col}'] = detected_row[col]
        
        results.append(result_row)
    
    # Crear DataFrame resultado
    result_df = pd.DataFrame(results)
    
    # Limpiar columnas duplicadas (mantener la primera)
    result_df = result_df.loc[:, ~result_df.columns.duplicated()].copy()
    
    # Renombrar columnas validated_ si no tienen prefijo
    column_mapping = {}
    for col in result_df.columns:
        if not col.startswith('validated_') and col not in ['candidato_tiempo_clean', 'nombre_archivo_normalized', 'is_canonical', 'has_match', 'num_matches', 'match_type', 'source_file']:
            if col == 'Folder':
                column_mapping[col] = 'validated_Folder'
            elif col == 'subfolder':
                column_mapping[col] = 'validated_subfolder'
            elif col == 'nombre_archivo':
                column_mapping[col] = 'validated_nombre_archivo'
            elif col == 'candidato tiempo':
                column_mapping[col] = 'validated_candidato_tiempo'
            elif col == 'SNR':
                column_mapping[col] = 'validated_SNR'
            elif col == 'choosen':
                column_mapping[col] = 'validated_choosen'
    
    if column_mapping:
        result_df = result_df.rename(columns=column_mapping)
    
    # Eliminar columnas auxiliares
    columns_to_drop = ['candidato_tiempo_clean', 'nombre_archivo_normalized', 'is_canonical']
    for col in columns_to_drop:
        if col in result_df.columns:
            result_df = result_df.drop(columns=[col])
    
    # Limpiar nuevamente después de renombrar
    result_df = result_df.loc[:, ~result_df.columns.duplicated()].copy()
    
    return result_df


def order_columns_by_relevance(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reordena las columnas del DataFrame por relevancia.
    
    Orden prioritario:
    1. Estado del Match: has_match, match_type
    2. Archivos: validated_nombre_archivo, detected_file
    3. Tiempos: validated_candidato_tiempo, detected_t_sec_dm_time
    4. Etiquetas/Clasificación: validated_choosen, detected_is_burst
    5. Métricas Clave: detected_snr_*, detected_detection_prob, detected_class_prob_*
    6. Información Complementaria: Resto de columnas
    """
    # Definir orden de columnas prioritarias
    priority_order = [
        # 1. Estado del Match
        'has_match',
        'match_type',
        'num_matches',
        'source_file',
        
        # 2. Archivos
        'validated_nombre_archivo',
        'detected_file',
        
        # 3. Tiempos
        'validated_candidato_tiempo',
        'detected_t_sec_dm_time',
        
        # 4. Etiquetas/Clasificación
        'validated_choosen',
        'detected_is_burst',
        'detected_is_burst_intensity',
        'detected_is_burst_linear',
        
        # 5. Métricas Clave
        'detected_snr_waterfall',
        'detected_snr_patch_dedispersed',
        'detected_snr_waterfall_linear',
        'detected_snr_patch_dedispersed_linear',
        'detected_detection_prob',
        'detected_class_prob_intensity',
        'detected_class_prob_linear',
        
        # 6. Validated adicionales
        'validated_Folder',
        'validated_subfolder',
        'validated_SNR',
    ]
    
    # Obtener todas las columnas del DataFrame
    all_columns = list(df.columns)
    
    # Separar columnas en prioritarias y complementarias
    priority_cols = [col for col in priority_order if col in all_columns]
    remaining_cols = [col for col in all_columns if col not in priority_order]
    
    # Ordenar columnas complementarias: primero validated_, luego detected_, luego otras
    validated_cols = [col for col in remaining_cols if col.startswith('validated_')]
    detected_cols = [col for col in remaining_cols if col.startswith('detected_')]
    other_cols = [col for col in remaining_cols if not col.startswith(('validated_', 'detected_'))]
    
    # Ordenar detected_cols por tipo de métrica
    detected_sorted = sorted(detected_cols)
    
    # Construir orden final
    final_order = priority_cols + validated_cols + detected_sorted + other_cols
    
    # Asegurar que todas las columnas estén incluidas
    missing_cols = [col for col in all_columns if col not in final_order]
    final_order.extend(missing_cols)
    
    return df[final_order]


def calculate_metrics(matches_df: pd.DataFrame, canonicos_df: pd.DataFrame, 
                      extras_df: pd.DataFrame) -> Dict:
    """
    Calcula métricas de rendimiento para una fase.
    
    Args:
        matches_df: DataFrame con todos los matches
        canonicos_df: DataFrame con canónicos validados
        extras_df: DataFrame con extras validados
        
    Returns:
        Diccionario con todas las métricas calculadas
    """
    metrics = {}
    
    # Limpiar columnas duplicadas (mantener la primera)
    matches_df = matches_df.loc[:, ~matches_df.columns.duplicated()].copy()
    
    # Separar matches por tipo
    validated_matches = matches_df[matches_df['match_type'] == 'validated'].copy().reset_index(drop=True)
    new_candidates = matches_df[matches_df['match_type'] == 'new'].copy().reset_index(drop=True)
    
    # Total de validados esperados
    total_validated = len(canonicos_df) + len(extras_df)
    
    # Validados con match - obtener validados únicos (sin duplicados)
    validated_with_match = validated_matches[validated_matches['has_match'] == True].copy()
    validated_with_match = validated_with_match.reset_index(drop=True)
    
    # Obtener validados únicos (un solo match por pulso validado)
    if len(validated_with_match) > 0:
        # Filtrar filas válidas manualmente
        valid_indices = []
        for idx in validated_with_match.index:
            if (pd.notna(validated_with_match.loc[idx, 'validated_nombre_archivo']) and 
                pd.notna(validated_with_match.loc[idx, 'validated_candidato_tiempo'])):
                valid_indices.append(idx)
        
        if len(valid_indices) > 0:
            valid_rows = validated_with_match.loc[valid_indices].copy()
            
            # Ordenar por probabilidad de detección (si existe) para quedarse con el mejor match
            if 'detected_detection_prob' in valid_rows.columns:
                valid_rows = valid_rows.sort_values('detected_detection_prob', ascending=False)
            
            # Obtener validados únicos - mantener el mejor match de cada pulso validado
            unique_validated = valid_rows.drop_duplicates(
                subset=['validated_nombre_archivo', 'validated_candidato_tiempo'],
                keep='first'
            )
            num_validated_matched = len(unique_validated)
        else:
            unique_validated = pd.DataFrame()
            num_validated_matched = 0
    else:
        unique_validated = pd.DataFrame()
        num_validated_matched = 0
    
    # Separar únicos en Canónicos y Extras usando identificador único (archivo + tiempo)
    unique_canonicos = pd.DataFrame()
    unique_extras = pd.DataFrame()
    
    if len(unique_validated) > 0:
        # Crear ID para filtrado robusto
        unique_validated['id_temp'] = unique_validated['validated_nombre_archivo'].astype(str) + '_' + unique_validated['validated_candidato_tiempo'].astype(str)
        canonicos_ids = canonicos_df['nombre_archivo'].astype(str) + '_' + canonicos_df['candidato tiempo'].astype(str)
        extras_ids = extras_df['nombre_archivo'].astype(str) + '_' + extras_df['candidato tiempo'].astype(str)
        
        unique_canonicos = unique_validated[unique_validated['id_temp'].isin(canonicos_ids)].copy()
        unique_extras = unique_validated[unique_validated['id_temp'].isin(extras_ids)].copy()
        
        # Limpiar columna temporal
        if 'id_temp' in unique_validated.columns:
            unique_validated = unique_validated.drop(columns=['id_temp'])
        if 'id_temp' in unique_canonicos.columns:
            unique_canonicos = unique_canonicos.drop(columns=['id_temp'])
        if 'id_temp' in unique_extras.columns:
            unique_extras = unique_extras.drop(columns=['id_temp'])
    
    num_canonicos_matched = len(unique_canonicos)
    num_extras_matched = len(unique_extras)
    
    # Métricas de clasificación DESGLOSADAS por clasificador (sobre validados únicos)
    # Canónicos - Clasificación Intensity
    num_canonicos_burst_intensity = 0
    if len(unique_canonicos) > 0:
        num_canonicos_burst_intensity = len(unique_canonicos[unique_canonicos['detected_is_burst_intensity'] == 'burst'])
    
    # Canónicos - Clasificación Linear
    num_canonicos_burst_linear = 0
    if len(unique_canonicos) > 0:
        num_canonicos_burst_linear = len(unique_canonicos[unique_canonicos['detected_is_burst_linear'] == 'burst'])
    
    # Canónicos - Clasificación Final (Pipeline)
    num_canonicos_burst_final = 0
    if len(unique_canonicos) > 0:
        num_canonicos_burst_final = len(unique_canonicos[unique_canonicos['detected_is_burst'] == 'burst'])
    
    # Extras - Clasificación Intensity
    num_extras_burst_intensity = 0
    if len(unique_extras) > 0:
        num_extras_burst_intensity = len(unique_extras[unique_extras['detected_is_burst_intensity'] == 'burst'])
    
    # Extras - Clasificación Linear
    num_extras_burst_linear = 0
    if len(unique_extras) > 0:
        num_extras_burst_linear = len(unique_extras[unique_extras['detected_is_burst_linear'] == 'burst'])
    
    # Extras - Clasificación Final (Pipeline)
    num_extras_burst_final = 0
    if len(unique_extras) > 0:
        num_extras_burst_final = len(unique_extras[unique_extras['detected_is_burst'] == 'burst'])
    
    # Totales (para compatibilidad con métricas globales)
    num_validated_burst_intensity = num_canonicos_burst_intensity + num_extras_burst_intensity
    num_validated_burst_linear = num_canonicos_burst_linear + num_extras_burst_linear
    num_validated_burst_final = num_canonicos_burst_final + num_extras_burst_final
    
    # Mantener compatibilidad con código anterior (usar final)
    num_canonicos_burst = num_canonicos_burst_final
    num_extras_burst = num_extras_burst_final
    num_validated_burst = num_validated_burst_final
    
    # Total de candidatos detectados
    total_detected = len(matches_df)
    num_new = len(new_candidates)
    
    # Métricas de clasificación BURST
    # Validados únicos con match clasificados como BURST
    if len(unique_validated) > 0:
        validated_burst = unique_validated[unique_validated['detected_is_burst'] == 'burst'].copy()
        num_validated_burst = len(validated_burst)
    else:
        num_validated_burst = 0
    
    # Nuevos clasificados - DESGLOSADO por clasificador
    new_burst_intensity = new_candidates[new_candidates['detected_is_burst_intensity'] == 'burst']
    new_burst_linear = new_candidates[new_candidates['detected_is_burst_linear'] == 'burst']
    new_burst_final = new_candidates[new_candidates['detected_is_burst'] == 'burst']
    new_no_burst_final = new_candidates[new_candidates['detected_is_burst'] == 'no_burst']
    
    # Mantener compatibilidad
    new_burst = new_burst_final
    new_no_burst = new_no_burst_final
    
    # Calcular métricas de clasificación
    # TP = validados detectados y clasificados como BURST
    # FN = validados no detectados
    # FP = nuevos clasificados como BURST
    # TN = nuevos clasificados como NO_BURST
    
    TP = num_validated_burst
    FN = total_validated - num_validated_matched
    FP = len(new_burst)
    TN = len(new_no_burst)
    
    # Recall (Sensibilidad) = TP / (TP + FN)
    recall = TP / (TP + FN) if (TP + FN) > 0 else 0.0
    
    # Precision = TP / (TP + FP)
    precision = TP / (TP + FP) if (TP + FP) > 0 else 0.0
    
    # Accuracy = (TP + TN) / Total
    total_classified = TP + FN + FP + TN
    accuracy = (TP + TN) / total_classified if total_classified > 0 else 0.0
    
    # F1-Score = 2 * (Precision * Recall) / (Precision + Recall)
    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0
    
    # Especificidad = TN / (TN + FP)
    specificity = TN / (TN + FP) if (TN + FP) > 0 else 0.0
    
    # Función auxiliar para calcular estadísticas de una columna
    def calculate_column_stats(df, col_name):
        """Calcula estadísticas descriptivas de una columna"""
        if col_name not in df.columns:
            return None
        values = pd.to_numeric(df[col_name], errors='coerce').dropna()
        if len(values) == 0:
            return None
        stats = {
            'mean': float(values.mean()),
            'median': float(values.median()),
            'min': float(values.min()),
            'max': float(values.max()),
            'count': len(values)
        }
        # Agregar std solo si hay más de 1 valor
        if len(values) > 1:
            stats['std'] = float(values.std())
        return stats
    
    # Métricas astronómicas para validados únicos con match (sin duplicados)
    validated_astronomical = {}
    if len(unique_validated) > 0:
        # SNR columns
        snr_cols = ['detected_snr_waterfall', 'detected_snr_patch_dedispersed', 
                   'detected_snr_waterfall_linear', 'detected_snr_patch_dedispersed_linear']
        for col in snr_cols:
            stats = calculate_column_stats(unique_validated, col)
            if stats:
                validated_astronomical[col] = stats
        
        # Probability columns
        prob_cols = ['detected_detection_prob', 'detected_class_prob_intensity', 
                    'detected_class_prob_linear']
        for col in prob_cols:
            stats = calculate_column_stats(unique_validated, col)
            if stats:
                validated_astronomical[col] = stats
        
        # Physical properties
        for col in ['detected_dm_pc_cm-3', 'detected_width_ms']:
            stats = calculate_column_stats(unique_validated, col)
            if stats:
                validated_astronomical[col] = stats
    
    # Métricas astronómicas DESGLOSADAS por clasificador para validados
    validated_by_classifier = {}
    if len(unique_validated) > 0:
        # Separar por clasificación Intensity
        validated_burst_intensity = unique_validated[unique_validated['detected_is_burst_intensity'] == 'burst'].copy()
        validated_no_burst_intensity = unique_validated[unique_validated['detected_is_burst_intensity'] != 'burst'].copy()
        
        # Separar por clasificación Linear
        validated_burst_linear = unique_validated[unique_validated['detected_is_burst_linear'] == 'burst'].copy()
        validated_no_burst_linear = unique_validated[unique_validated['detected_is_burst_linear'] != 'burst'].copy()
        
        validated_by_classifier = {
            'intensity': {
                'burst': {
                    'count': len(validated_burst_intensity),
                    'class_prob_stats': calculate_column_stats(validated_burst_intensity, 'detected_class_prob_intensity'),
                    'snr_patch_stats': calculate_column_stats(validated_burst_intensity, 'detected_snr_patch_dedispersed')
                },
                'no_burst': {
                    'count': len(validated_no_burst_intensity),
                    'class_prob_stats': calculate_column_stats(validated_no_burst_intensity, 'detected_class_prob_intensity'),
                    'snr_patch_stats': calculate_column_stats(validated_no_burst_intensity, 'detected_snr_patch_dedispersed')
                }
            },
            'linear': {
                'burst': {
                    'count': len(validated_burst_linear),
                    'class_prob_stats': calculate_column_stats(validated_burst_linear, 'detected_class_prob_linear'),
                    'snr_linear_stats': calculate_column_stats(validated_burst_linear, 'detected_snr_patch_dedispersed_linear')
                },
                'no_burst': {
                    'count': len(validated_no_burst_linear),
                    'class_prob_stats': calculate_column_stats(validated_no_burst_linear, 'detected_class_prob_linear'),
                    'snr_linear_stats': calculate_column_stats(validated_no_burst_linear, 'detected_snr_patch_dedispersed_linear')
                }
            }
        }
    
    # Métricas astronómicas para nuevos
    new_astronomical = {}
    if len(new_candidates) > 0:
        snr_cols = ['detected_snr_waterfall', 'detected_snr_patch_dedispersed', 
                   'detected_snr_waterfall_linear', 'detected_snr_patch_dedispersed_linear']
        for col in snr_cols:
            if col in new_candidates.columns:
                values = pd.to_numeric(new_candidates[col], errors='coerce').dropna()
                if len(values) > 0:
                    new_astronomical[col] = {
                        'mean': float(values.mean()),
                        'median': float(values.median()),
                        'min': float(values.min()),
                        'max': float(values.max())
                    }
    
    # Coherencia I+L para validados únicos
    validated_coherence = {}
    if len(unique_validated) > 0:
        both_burst = unique_validated[
            (unique_validated['detected_is_burst_intensity'] == 'burst') &
            (unique_validated['detected_is_burst_linear'] == 'burst')
        ]
        only_intensity = unique_validated[
            (unique_validated['detected_is_burst_intensity'] == 'burst') &
            (unique_validated['detected_is_burst_linear'] != 'burst')
        ]
        only_linear = unique_validated[
            (unique_validated['detected_is_burst_intensity'] != 'burst') &
            (unique_validated['detected_is_burst_linear'] == 'burst')
        ]
        neither = unique_validated[
            (unique_validated['detected_is_burst_intensity'] != 'burst') &
            (unique_validated['detected_is_burst_linear'] != 'burst')
        ]
        
        validated_coherence = {
            'both_burst': len(both_burst),
            'only_intensity': len(only_intensity),
            'only_linear': len(only_linear),
            'neither': len(neither)
        }
    
    metrics = {
        'total_validated': total_validated,
        'num_canonicos': len(canonicos_df),
        'num_extras': len(extras_df),
        'num_canonicos_matched': num_canonicos_matched,
        'num_extras_matched': num_extras_matched,
        'num_validated_matched': num_validated_matched,
        # Métricas desglosadas por clasificador - Canónicos
        'num_canonicos_burst_intensity': num_canonicos_burst_intensity,
        'num_canonicos_burst_linear': num_canonicos_burst_linear,
        'num_canonicos_burst_final': num_canonicos_burst_final,
        # Métricas desglosadas por clasificador - Extras
        'num_extras_burst_intensity': num_extras_burst_intensity,
        'num_extras_burst_linear': num_extras_burst_linear,
        'num_extras_burst_final': num_extras_burst_final,
        # Métricas desglosadas por clasificador - Totales
        'num_validated_burst_intensity': num_validated_burst_intensity,
        'num_validated_burst_linear': num_validated_burst_linear,
        'num_validated_burst_final': num_validated_burst_final,
        # Compatibilidad (usar final)
        'num_canonicos_burst': num_canonicos_burst_final,
        'num_extras_burst': num_extras_burst_final,
        'num_validated_burst': num_validated_burst_final,
        # Nuevos desglosados
        'num_new_burst_intensity': len(new_burst_intensity),
        'num_new_burst_linear': len(new_burst_linear),
        'num_new_burst_final': len(new_burst_final),
        'num_new_no_burst_final': len(new_no_burst_final),
        # Compatibilidad
        'num_new_burst': len(new_burst_final),
        'num_new_no_burst': len(new_no_burst_final),
        # Métricas globales
        'total_detected': total_detected,
        'num_new': num_new,
        'recall': recall,
        'precision': precision,
        'accuracy': accuracy,
        'f1_score': f1_score,
        'specificity': specificity,
        # Estadísticas astronómicas
        'validated_astronomical': validated_astronomical,
        'new_astronomical': new_astronomical,
        'validated_by_classifier': validated_by_classifier,
        'validated_coherence': validated_coherence,
        'TP': TP,
        'FN': FN,
        'FP': FP,
        'TN': TN
    }
    
    return metrics


def save_to_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Guarda el DataFrame en formato Excel con formato mejorado y columnas ordenadas.
    
    Args:
        df: DataFrame a guardar
        output_path: Ruta del archivo Excel de salida
    """
    print(f"\nGuardando resultados en: {output_path}")
    
    # Reordenar columnas por relevancia
    df_ordered = order_columns_by_relevance(df)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
            df_ordered.to_excel(writer, index=False, sheet_name='Matches')
            
            worksheet = writer.sheets['Matches']
            
            # Formatear encabezado
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df_ordered.columns, start=1):
                try:
                    col_max = df_ordered[col].astype(str).map(len).max()
                    if pd.isna(col_max):
                        col_max = 0
                    max_length = max(int(col_max), len(str(col))) + 2
                    max_length = min(max_length, 50)
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = max_length
                except Exception:
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = 15
            
            # Congelar primera fila
            worksheet.freeze_panes = 'A2'
        
        print(f"  [OK] Archivo Excel guardado exitosamente")
        
    except PermissionError as e:
        print(f"  [ERROR] No se puede escribir el archivo. Por favor, cierra el archivo Excel si está abierto.")
        print(f"  Error: {e}")
        raise
    except ImportError:
        print("⚠ openpyxl no está instalado. Instalando...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        save_to_excel(df, output_path)


def get_phase_config(config: Dict, phase_name: str) -> Dict:
    """
    Obtiene la configuración de fases para un caso específico.
    
    Args:
        config: Configuración completa
        phase_name: Nombre del caso (ej: 'all', 'no-phase2', etc.)
        
    Returns:
        Diccionario con configuración de fases
    """
    hf_config = config.get('high_frequency', {})
    
    # Mapeo de nombres de casos a configuraciones
    phase_configs = {
        'all': {
            'phase2': True,
            'phase3a': True,
            'phase3b': True
        },
        'no-phase2': {
            'phase2': False,
            'phase3a': True,
            'phase3b': True
        },
        'no-class-intensity': {
            'phase2': hf_config.get('enable_linear_validation', False),
            'phase3a': False,
            'phase3b': hf_config.get('enable_linear_classification', False)
        },
        'no-classification-linear': {
            'phase2': hf_config.get('enable_linear_validation', False),
            'phase3a': True,
            'phase3b': False
        },
        'no-phase2-no-classification-intensity': {
            'phase2': False,
            'phase3a': False,
            'phase3b': hf_config.get('enable_linear_classification', False)
        },
        'no-phase2-no-classification-linear': {
            'phase2': False,
            'phase3a': True,
            'phase3b': False
        }
    }
    
    return phase_configs.get(phase_name, {})


def main():
    """Función principal del script."""
    base_path = Path("ResultsThesis/ALMA-4phases")
    config_path = Path("config.yaml")
    
    # Cargar configuración
    config = load_config(str(config_path))
    
    # Archivo de candidatos validados
    validated_path = base_path / "tabla_transformada.xlsx"
    
    # Archivos de detecciones a comparar
    detection_files = [
        ("combined_candidates-all.csv", "matches_all.xlsx", "all"),
        ("combined_candidates-no-class-intensity.csv", "matches_no-class-intensity.xlsx", "no-class-intensity"),
        ("combined_candidates-no-classification-linear.csv", "matches_no-classification-linear.xlsx", "no-classification-linear"),
        ("combined_candidates-no-phase2-no-classification-intensity.csv", 
         "matches_no-phase2-no-classification-intensity.xlsx", "no-phase2-no-classification-intensity"),
        ("combined_candidates-no-phase2-no-classification-linear.csv", 
         "matches_no-phase2-no-classification-linear.xlsx", "no-phase2-no-classification-linear"),
        ("combined_candidates-no-phase2.csv", "matches_no-phase2.xlsx", "no-phase2"),
    ]
    
    # Cargar candidatos validados
    canonicos_df, extras_df = load_validated_candidates(str(validated_path))
    all_validated_df = pd.concat([canonicos_df, extras_df], ignore_index=True)
    
    # Diccionario para almacenar métricas de todas las fases
    all_metrics = {}
    
    # Procesar cada archivo de detecciones
    for detection_file, output_file, phase_name in detection_files:
        detection_path = base_path / detection_file
        
        if not detection_path.exists():
            print(f"\n⚠ Advertencia: No se encontró {detection_path}")
            continue
        
        print(f"\n{'='*60}")
        print(f"Procesando: {detection_file} (Fase: {phase_name})")
        print(f"{'='*60}")
        
        # Cargar candidatos detectados
        detected_df = load_detected_candidates(str(detection_path))
        
        # Encontrar matches
        matches_df = find_matches(all_validated_df, detected_df, detection_file)
        
        # Calcular métricas
        metrics = calculate_metrics(matches_df, canonicos_df, extras_df)
        all_metrics[phase_name] = metrics
        
        # Guardar en Excel
        output_path = base_path / output_file
        try:
            save_to_excel(matches_df, str(output_path))
        except PermissionError:
            print(f"  ⚠ Saltando {output_file} - archivo está abierto. Por favor ciérralo y vuelve a ejecutar.")
            continue
        
        # Mostrar resumen
        print(f"\n  Resumen de métricas:")
        print(f"    Canónicos detectados: {metrics['num_canonicos_matched']}/{metrics['num_canonicos']}")
        print(f"    Extras detectados: {metrics['num_extras_matched']}/{metrics['num_extras']}")
        print(f"    Total validados detectados: {metrics['num_validated_matched']}/{metrics['total_validated']}")
        print(f"    Candidatos nuevos: {metrics['num_new']}")
        print(f"    Recall: {metrics['recall']:.3f}")
        print(f"    Precision: {metrics['precision']:.3f}")
        print(f"    F1-Score: {metrics['f1_score']:.3f}")
        print(f"    Accuracy: {metrics['accuracy']:.3f}")
    
    # Guardar métricas en JSON para el reporte
    import json
    metrics_path = base_path / "metrics_all_phases.json"
    with open(metrics_path, 'w', encoding='utf-8') as f:
        json.dump(all_metrics, f, indent=2, ensure_ascii=False)
    print(f"\n{'='*60}")
    print(f"[OK] Métricas guardadas en: {metrics_path}")
    print(f"[OK] Proceso completado!")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()

