#!/usr/bin/env python3
"""
Script para comparar candidatos validados manualmente con detecciones automáticas.

Compara tabla_transformada.csv con múltiples archivos combined_candidates-*.csv
y genera archivos Excel con todos los matches encontrados.
"""

import re
import pandas as pd
from pathlib import Path
from typing import Tuple, List
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def filter_special_files(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra archivos con formato especial que contienen tiempos en el nombre.
    Estos archivos ya están cubiertos por los archivos normales.
    
    EXCEPCIÓN: Si un archivo especial es el único que contiene un canónico
    (ej: 242_0005_t44.169), se mantiene porque no existe archivo normal.
    
    Args:
        df: DataFrame con candidatos detectados
        
    Returns:
        DataFrame filtrado sin archivos especiales (excepto excepciones)
    """
    if 'file' not in df.columns:
        return df
    
    # Patrón: archivos que contienen _t seguido de números (ej: _t10.882, _t2.3_t17.395)
    pattern = r'_t[\d.]+'
    special_mask = df['file'].astype(str).str.contains(pattern, regex=True, na=False)
    
    # EXCEPCIÓN: Archivo 242_0005_t44.169 - este canónico solo existe en archivo especial
    # Verificar si existe archivo normal para 242_0005
    normal_242_pattern = r'2017-04-03-13_38_31_0005\.fits'
    has_normal_242 = df['file'].astype(str).str.contains(normal_242_pattern, regex=True, na=False).any()
    
    # Si NO existe archivo normal, mantener el especial
    if not has_normal_242:
        # Mantener archivo especial 242_0005_t44.169
        exception_pattern = r'13_38_31.*242_0005.*_t44\.169'
        exception_mask = df['file'].astype(str).str.contains(exception_pattern, regex=True, na=False)
        special_mask = special_mask & ~exception_mask
        if exception_mask.any():
            print(f"  [INFO] Excepción: Manteniendo archivo especial 242_0005_t44.169 (no existe archivo normal)")
    
    filtered_count = special_mask.sum()
    if filtered_count > 0:
        print(f"  [INFO] Filtrados {filtered_count} archivos especiales con tiempos en el nombre")
    
    return df[~special_mask]


def normalize_filename(filename: str) -> str:
    """
    Normaliza un nombre de archivo para comparación robusta.
    Maneja variaciones: guiones, guiones bajos, formatos compactos.
    
    Ejemplos:
    - "2017-04-03-13-38-31_0006" → "2017_04_03_13_38_31_0006"
    - "2017-04-03-13_38_31_0006" → "2017_04_03_13_38_31_0006"
    - "2017-04-03-133831_0006" → "2017_04_03_13_38_31_0006"
    - "2017-04-03-134701_0004" → "2017_04_03_13_47_01_0004"
    
    Args:
        filename: Nombre de archivo a normalizar
        
    Returns:
        Nombre de archivo normalizado en formato: YYYY_MM_DD_HH_MM_SS_XXXX
    """
    if pd.isna(filename) or filename == '':
        return ''
    
    filename = str(filename).strip()
    
    # Quitar extensión .fits
    filename = filename.replace('.fits', '')
    
    # CASO ESPECIAL: Archivos con formato _tXX.XXX (archivos especiales)
    # Ejemplo: "2017-04-03-13_38_31_242_0005_t44.169"
    # Extraer el nombre base antes del _t para normalización
    special_match = re.match(r'(.+?)_t[\d.]+', filename)
    if special_match:
        filename = special_match.group(1)  # Tomar solo la parte antes de _t
    
    # Convertir a minúsculas
    filename = filename.lower()
    
    # Patrón para extraer componentes:
    # Formato esperado: YYYY-MM-DD-HH-MM-SS_XXXX o variaciones
    # También puede ser: YYYY-MM-DD-HHMMSS_XXXX (compacto)
    
    # Intentar extraer fecha (YYYY-MM-DD o YYYY_MM_DD)
    date_match = re.match(r'(\d{4})[-_](\d{2})[-_](\d{2})', filename)
    if not date_match:
        # Si no hay match, intentar formato más flexible
        # Fallback: convertir todos los guiones a guiones bajos
        filename = filename.replace('-', '_')
        return filename
    
    year = date_match.group(1)
    month = date_match.group(2)
    day = date_match.group(3)
    
    # Resto después de la fecha
    remaining = filename[date_match.end():]
    
    # Extraer subfolder (últimos 4 dígitos al final)
    subfolder_match = re.search(r'_(\d{4})$', remaining)
    if subfolder_match:
        subfolder = subfolder_match.group(1)
        # Remover subfolder del remaining para procesar hora
        remaining = remaining[:subfolder_match.start()]
    else:
        # Intentar sin guion bajo antes del subfolder
        subfolder_match = re.search(r'(\d{4})$', remaining)
        if subfolder_match:
            subfolder = subfolder_match.group(1)
            remaining = remaining[:subfolder_match.start()]
        else:
            subfolder = ''
    
    # Procesar hora: puede ser HH-MM-SS, HH_MM_SS, o HHMMSS
    # Remover cualquier separador inicial
    remaining = remaining.lstrip('-_')
    
    # Intentar extraer hora con separadores
    time_match_sep = re.match(r'(\d{2})[-_](\d{2})[-_](\d{2})', remaining)
    if time_match_sep:
        hour = time_match_sep.group(1)
        minute = time_match_sep.group(2)
        second = time_match_sep.group(3)
    else:
        # Intentar formato compacto (HHMMSS - 6 dígitos)
        time_match_compact = re.match(r'(\d{6})', remaining)
        if time_match_compact:
            time_str = time_match_compact.group(1)
            hour = time_str[0:2]
            minute = time_str[2:4]
            second = time_str[4:6]
        else:
            # Si no se puede extraer, usar el remaining tal cual
            # Esto puede pasar si hay formato inesperado
            hour = minute = second = ''
    
    # Reconstruir en formato normalizado
    if hour and minute and second:
        normalized = f"{year}_{month}_{day}_{hour}_{minute}_{second}"
    else:
        # Fallback: usar formato original sin separadores de hora
        normalized = f"{year}_{month}_{day}"
        if remaining:
            # Agregar remaining sin separadores problemáticos
            normalized += "_" + remaining.replace('-', '_').replace('_', '')
    
    if subfolder:
        normalized += f"_{subfolder}"
    
    return normalized


def load_validated_candidates(csv_path: str) -> pd.DataFrame:
    """
    Carga y normaliza los candidatos validados manualmente.
    
    Args:
        csv_path: Ruta al archivo tabla_transformada.csv
        
    Returns:
        DataFrame con candidatos validados normalizados
    """
    print(f"Cargando candidatos validados desde: {csv_path}")
    df = pd.read_csv(csv_path, sep='\t', encoding='utf-8')
    
    # Normalizar nombres de columnas
    df.columns = df.columns.str.strip()
    
    # Normalizar campo 'nombre_archivo' para matching consistente
    if 'nombre_archivo' in df.columns:
        df['nombre_archivo_normalized'] = df['nombre_archivo'].apply(normalize_filename)
    else:
        print("⚠ Advertencia: No se encontró la columna 'nombre_archivo'")
        df['nombre_archivo_normalized'] = ''
    
    # Normalizar campo 'candidato tiempo' - remover texto entre paréntesis
    if 'candidato tiempo' in df.columns:
        df['candidato_tiempo_clean'] = df['candidato tiempo'].astype(str).apply(
            lambda x: re.sub(r'\s*\([^)]*\)', '', x).strip()
        )
        # Convertir a numérico
        df['candidato_tiempo_clean'] = pd.to_numeric(
            df['candidato_tiempo_clean'], 
            errors='coerce'
        )
    else:
        print("⚠ Advertencia: No se encontró la columna 'candidato tiempo'")
        df['candidato_tiempo_clean'] = None
    
    print(f"  [OK] Cargados {len(df)} candidatos validados")
    return df


def load_detected_candidates(csv_path: str) -> pd.DataFrame:
    """
    Carga y normaliza los candidatos detectados automáticamente.
    
    Args:
        csv_path: Ruta al archivo combined_candidates-*.csv
        
    Returns:
        DataFrame con candidatos detectados normalizados
    """
    print(f"Cargando candidatos detectados desde: {csv_path}")
    df = pd.read_csv(csv_path, encoding='utf-8')
    
    # Filtrar archivos especiales con tiempos en el nombre
    df = filter_special_files(df)
    
    # Normalizar campo 'file' para matching consistente
    if 'file' in df.columns:
        df['file_clean'] = df['file'].apply(normalize_filename)
    else:
        print("⚠ Advertencia: No se encontró la columna 'file'")
        df['file_clean'] = ''
    
    # Asegurar que t_sec_dm_time sea numérico
    if 't_sec_dm_time' in df.columns:
        df['t_sec_dm_time'] = pd.to_numeric(df['t_sec_dm_time'], errors='coerce')
    else:
        print("⚠ Advertencia: No se encontró la columna 't_sec_dm_time'")
        df['t_sec_dm_time'] = None
    
    print(f"  [OK] Cargados {len(df)} candidatos detectados (después de filtrar archivos especiales)")
    return df


def find_matches(validated_df: pd.DataFrame, detected_df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    """
    Encuentra matches entre candidatos validados y detectados.
    También incluye candidatos detectados que no tienen match (etiquetados como "new").
    
    Args:
        validated_df: DataFrame con candidatos validados
        detected_df: DataFrame con candidatos detectados
        source_name: Nombre del archivo fuente (para la columna source_file)
        
    Returns:
        DataFrame con todos los matches y candidatos nuevos
    """
    print(f"\nBuscando matches...")
    
    results = []
    matched_detected_indices = set()  # Para rastrear qué detecciones ya fueron matcheadas
    
    # Primero: procesar todos los candidatos validados
    for idx, validated_row in validated_df.iterrows():
        nombre_archivo_normalized = validated_row.get('nombre_archivo_normalized', '')
        candidato_tiempo = validated_row.get('candidato_tiempo_clean')
        
        if pd.isna(nombre_archivo_normalized) or nombre_archivo_normalized == '' or pd.isna(candidato_tiempo):
            # Si falta información crítica, crear fila sin match
            result_row = validated_row.to_dict()
            result_row['has_match'] = False
            result_row['num_matches'] = 0
            result_row['source_file'] = source_name
            result_row['match_type'] = 'validated'  # Tipo: candidato validado sin match
            # Agregar columnas detected_ vacías
            for col in detected_df.columns:
                if col not in ['file_clean']:  # Excluir columna auxiliar
                    result_row[f'detected_{col}'] = None
            results.append(result_row)
            continue
        
        # Filtrar por archivo usando nombres normalizados
        file_matches = detected_df[
            detected_df['file_clean'] == nombre_archivo_normalized
        ].copy()
        
        if len(file_matches) == 0:
            # No hay matches por archivo
            result_row = validated_row.to_dict()
            result_row['has_match'] = False
            result_row['num_matches'] = 0
            result_row['source_file'] = source_name
            result_row['match_type'] = 'validated'  # Tipo: candidato validado sin match
            # Agregar columnas detected_ vacías
            for col in detected_df.columns:
                if col not in ['file_clean']:
                    result_row[f'detected_{col}'] = None
            results.append(result_row)
            continue
        
        # Filtrar por tiempo (diferencia < 0.1 segundos)
        file_matches['time_diff'] = abs(file_matches['t_sec_dm_time'] - candidato_tiempo)
        time_matches = file_matches[file_matches['time_diff'] < 0.1].copy()
        
        num_matches = len(time_matches)
        
        if num_matches == 0:
            # No hay matches por tiempo
            result_row = validated_row.to_dict()
            result_row['has_match'] = False
            result_row['num_matches'] = 0
            result_row['source_file'] = source_name
            result_row['match_type'] = 'validated'  # Tipo: candidato validado sin match
            # Agregar columnas detected_ vacías
            for col in detected_df.columns:
                if col not in ['file_clean', 'time_diff']:
                    result_row[f'detected_{col}'] = None
            results.append(result_row)
        else:
            # Hay matches - crear una fila por cada match
            for match_idx, match_row in time_matches.iterrows():
                matched_detected_indices.add(match_idx)  # Marcar como matcheado
                result_row = validated_row.to_dict()
                result_row['has_match'] = True
                result_row['num_matches'] = num_matches
                result_row['source_file'] = source_name
                result_row['match_type'] = 'validated'  # Tipo: candidato validado
                # Agregar todas las columnas detected_
                for col in detected_df.columns:
                    if col not in ['file_clean', 'time_diff']:
                        result_row[f'detected_{col}'] = match_row[col]
                results.append(result_row)
    
    # Segundo: agregar candidatos detectados que NO tienen match (etiquetados como "new")
    print(f"\nBuscando candidatos detectados sin match...")
    unmatched_detected = detected_df[~detected_df.index.isin(matched_detected_indices)].copy()
    
    for idx, detected_row in unmatched_detected.iterrows():
        result_row = {}
        # Columnas validated_ vacías
        for col in validated_df.columns:
            if col not in ['candidato_tiempo_clean']:
                if col.startswith('validated_'):
                    result_row[col] = None
                elif col == 'Folder':
                    result_row['validated_Folder'] = None
                elif col == 'subfolder':
                    result_row['validated_subfolder'] = None
                elif col == 'nombre_archivo':
                    result_row['validated_nombre_archivo'] = None
                elif col == 'candidato tiempo':
                    result_row['validated_candidato_tiempo'] = None
                elif col == 'SNR':
                    result_row['validated_SNR'] = None
                elif col == 'choosen':
                    result_row['validated_choosen'] = None
                else:
                    result_row[f'validated_{col}'] = None
        
        result_row['has_match'] = False
        result_row['num_matches'] = 0
        result_row['source_file'] = source_name
        result_row['match_type'] = 'new'  # Tipo: candidato nuevo (no validado)
        
        # Agregar todas las columnas detected_
        for col in detected_df.columns:
            if col not in ['file_clean', 'time_diff']:
                result_row[f'detected_{col}'] = detected_row[col]
        
        results.append(result_row)
    
    print(f"  [OK] Candidatos nuevos (sin match): {len(unmatched_detected)}")
    
    # Crear DataFrame resultado
    result_df = pd.DataFrame(results)
    
    # Reordenar columnas: primero validated_, luego detected_, luego adicionales
    validated_cols = [c for c in result_df.columns if c.startswith('validated_') or 
                     c in ['Folder', 'subfolder', 'nombre_archivo', 'candidato tiempo', 
                           'SNR', 'choosen']]
    detected_cols = [c for c in result_df.columns if c.startswith('detected_')]
    additional_cols = ['has_match', 'num_matches', 'match_type', 'source_file']
    
    # Eliminar columnas auxiliares si existen
    columns_to_drop = ['candidato_tiempo_clean', 'nombre_archivo_normalized']
    for col in columns_to_drop:
        if col in result_df.columns:
            result_df = result_df.drop(columns=[col])
    
    # Renombrar columnas validated_ si no tienen prefijo
    column_mapping = {}
    for col in validated_cols:
        if not col.startswith('validated_') and col != 'candidato_tiempo_clean':
            if col == 'Folder':
                column_mapping[col] = 'validated_Folder'
            elif col == 'subfolder':
                column_mapping[col] = 'validated_subfolder'
            elif col == 'nombre_archivo':
                column_mapping[col] = 'validated_nombre_archivo'
            elif col == 'candidato tiempo':
                column_mapping[col] = 'validated_candidato_tiempo'
            elif col == 'SNR':
                column_mapping[col] = 'validated_SNR'
            elif col == 'choosen':
                column_mapping[col] = 'validated_choosen'
    
    if column_mapping:
        result_df = result_df.rename(columns=column_mapping)
        validated_cols = [column_mapping.get(c, c) for c in validated_cols]
    
    # Reordenar
    all_cols = validated_cols + detected_cols + additional_cols
    existing_cols = [c for c in all_cols if c in result_df.columns]
    result_df = result_df[existing_cols]
    
    print(f"  [OK] Encontrados {len(result_df)} filas de resultados")
    print(f"  [OK] Matches encontrados: {result_df['has_match'].sum()}")
    print(f"  [OK] Candidatos validados sin match: {len(result_df[(result_df['match_type'] == 'validated') & (~result_df['has_match'])])}")
    print(f"  [OK] Candidatos nuevos (detectados sin validar): {len(result_df[result_df['match_type'] == 'new'])}")
    
    return result_df


def save_to_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Guarda el DataFrame en formato Excel con formato mejorado.
    
    Args:
        df: DataFrame a guardar
        output_path: Ruta del archivo Excel de salida
    """
    print(f"\nGuardando resultados en: {output_path}")
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name='Matches')
            
            # Obtener la hoja de trabajo
            worksheet = writer.sheets['Matches']
            
            # Formatear encabezado
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns, start=1):
                try:
                    col_max = df[col].astype(str).map(len).max()
                    if pd.isna(col_max):
                        col_max = 0
                    max_length = max(int(col_max), len(str(col))) + 2
                    max_length = min(max_length, 50)
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = max_length
                except Exception:
                    # Si hay error, usar ancho por defecto
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = 15
            
            # Congelar primera fila
            worksheet.freeze_panes = 'A2'
        
        print(f"  [OK] Archivo Excel guardado exitosamente")
        
    except PermissionError as e:
        print(f"  [ERROR] No se puede escribir el archivo. Por favor, cierra el archivo Excel si está abierto.")
        print(f"  Error: {e}")
        raise
    except ImportError:
        print("⚠ openpyxl no está instalado. Instalando...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        # Reintentar
        save_to_excel(df, output_path)


def main():
    """Función principal del script."""
    base_path = Path("ResultsThesis")
    
    # Archivo de candidatos validados
    validated_path = base_path / "tabla_transformada.csv"
    
    # Archivos de detecciones a comparar
    detection_files = [
        ("combined_candidates-all.csv", "matches_all.xlsx"),
        ("combined_candidates-no-class-intensity.csv", "matches_no-class-intensity.xlsx"),
        ("combined_candidates-no-classification-linear.csv", "matches_no-classification-linear.xlsx"),
        ("combined_candidates-no-phase2.csv", "matches_no-phase2.xlsx"),
        ("combined_candidates-no-phase2-no-classification-intensity.csv", 
         "matches_no-phase2-no-classification-intensity.xlsx"),
        ("combined_candidates-no-phase2-no-classification-linear.csv", 
         "matches_no-phase2-no-classification-linear.xlsx"),
    ]
    
    # Cargar candidatos validados
    validated_df = load_validated_candidates(str(validated_path))
    
    # Procesar cada archivo de detecciones
    for detection_file, output_file in detection_files:
        detection_path = base_path / detection_file
        
        if not detection_path.exists():
            print(f"\n⚠ Advertencia: No se encontró {detection_path}")
            continue
        
        print(f"\n{'='*60}")
        print(f"Procesando: {detection_file}")
        print(f"{'='*60}")
        
        # Cargar candidatos detectados
        detected_df = load_detected_candidates(str(detection_path))
        
        # Encontrar matches
        matches_df = find_matches(validated_df, detected_df, detection_file)
        
        # Guardar en Excel
        output_path = base_path / output_file
        try:
            save_to_excel(matches_df, str(output_path))
        except PermissionError:
            print(f"  ⚠ Saltando {output_file} - archivo está abierto. Por favor ciérralo y vuelve a ejecutar.")
            continue
    
    print(f"\n{'='*60}")
    print("[OK] Proceso completado!")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()

