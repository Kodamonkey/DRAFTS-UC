#!/usr/bin/env python3
"""
Script para transformar una tabla agregando el nombre real del archivo basado en config.yaml.

El script:
1. Lee el config.yaml para mapear Folder + subfolder -> nombre_archivo
2. Lee la tabla de entrada (CSV o texto)
3. Agrega una columna con el nombre real del archivo
4. Guarda la tabla transformada
"""

import re
import sys
import pandas as pd
from pathlib import Path
from typing import Dict, Tuple, Optional


def parse_config_yaml(config_path: str) -> Dict[Tuple[str, str], str]:
    """
    Parsea el config.yaml y crea un mapeo de (Folder, subfolder) -> nombre_archivo.
    
    Args:
        config_path: Ruta al archivo config.yaml
        
    Returns:
        Diccionario con clave (folder, subfolder) y valor nombre_archivo
    """
    mapping = {}
    current_folder = None
    
    with open(config_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    for line in lines:
        stripped = line.strip()
        
        # Buscar comentarios con formato # NoXXXX (puede tener espacios antes)
        folder_match = re.search(r'#\s*(No\d+)', stripped)
        if folder_match:
            current_folder = folder_match.group(1)
            continue
        
        # Buscar líneas con archivos (formato: - "nombre_archivo")
        if stripped.startswith('- "') and current_folder:
            # Extraer el nombre del archivo
            archivo_match = re.search(r'-\s*"([^"]+)"', stripped)
            if archivo_match:
                nombre_archivo = archivo_match.group(1)
                
                # Extraer el subfolder de los últimos 4 números después del último guion bajo
                # Ejemplos:
                # "2017-04-03-08_16_13_0006" -> subfolder = "0006" o "6"
                # "2017-04-03-08_16_13_142_0006_t10.882_t25.829" -> subfolder = "0006" o "6"
                # "2017-04-03-133831_0008" -> subfolder = "0008" o "8"
                subfolder_match = re.search(r'_(\d{4})(?=[^0-9]|$)', nombre_archivo)
                if not subfolder_match:
                    # Intentar buscar los últimos 4 dígitos del string
                    subfolder_match = re.search(r'(\d{4})$', nombre_archivo)
                
                if subfolder_match:
                    subfolder_full = subfolder_match.group(1)  # "0006"
                    subfolder_int = str(int(subfolder_full))  # "6"
                    
                    # Mapear tanto con el formato completo como con el número
                    mapping[(current_folder, subfolder_int)] = nombre_archivo
                    mapping[(current_folder, subfolder_full)] = nombre_archivo
                    mapping[(current_folder, f"_{subfolder_full}")] = nombre_archivo
                    
                    # También mapear variantes del folder (con y sin _no)
                    # Esto permite que "No0134_no" en la tabla mapee a "No0134" en el config
                    mapping[(f"{current_folder}_no", subfolder_int)] = nombre_archivo
                    mapping[(f"{current_folder}_no", subfolder_full)] = nombre_archivo
    
    return mapping


def transform_table(input_file: str, config_path: str, output_file: Optional[str] = None) -> pd.DataFrame:
    """
    Transforma la tabla agregando el nombre real del archivo.
    
    Args:
        input_file: Ruta al archivo de entrada (CSV o texto con tabs)
        config_path: Ruta al config.yaml
        output_file: Ruta opcional para guardar el resultado
        
    Returns:
        DataFrame con la columna adicional
    """
    # Leer el mapeo del config.yaml
    print(f"Leyendo config.yaml desde: {config_path}")
    mapping = parse_config_yaml(config_path)
    print(f"Encontrados {len(mapping)} mapeos")
    
    # Leer la tabla de entrada
    print(f"\nLeyendo tabla desde: {input_file}")
    
    # Intentar leer como CSV primero, si no funciona, leer como texto con tabs
    try:
        df = pd.read_csv(input_file, sep='\t', encoding='utf-8')
    except:
        try:
            df = pd.read_csv(input_file, sep=',', encoding='utf-8')
        except:
            # Leer como texto plano y parsear manualmente
            with open(input_file, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f.readlines() if line.strip()]
            
            # Asumir que la primera línea es el encabezado
            headers = lines[0].split('\t') if '\t' in lines[0] else lines[0].split()
            data = []
            for line in lines[1:]:
                row = line.split('\t') if '\t' in line else line.split()
                if len(row) == len(headers):
                    data.append(row)
            
            df = pd.DataFrame(data, columns=headers)
    
    print(f"Filas leídas: {len(df)}")
    print(f"Columnas: {list(df.columns)}")
    
    # Normalizar nombres de columnas (eliminar espacios)
    df.columns = df.columns.str.strip()
    
    # Función para encontrar el nombre del archivo
    def get_filename(row):
        folder = str(row.get('Folder', '')).strip()
        subfolder = str(row.get('subfolder', '')).strip()
        
        # Limpiar subfolder (puede tener texto adicional como "(dudosos)")
        subfolder_clean = re.sub(r'\([^)]*\)', '', subfolder).strip()
        
        # Intentar diferentes variaciones del subfolder
        keys_to_try = [
            (folder, subfolder_clean),
            (folder, subfolder),
            (folder, subfolder_clean.lstrip('0') if subfolder_clean.isdigit() else subfolder_clean),
            (folder, f"{int(subfolder_clean):04d}" if subfolder_clean.isdigit() else subfolder_clean),
        ]
        
        for key in keys_to_try:
            if key in mapping:
                return mapping[key]
        
        # Si no se encuentra, retornar None
        return None
    
    # Agregar la columna con el nombre del archivo
    print("\nAgregando columna 'nombre_archivo'...")
    df['nombre_archivo'] = df.apply(get_filename, axis=1)
    
    # Reordenar columnas: Folder, subfolder, nombre_archivo, y luego el resto
    cols = ['Folder', 'subfolder', 'nombre_archivo']
    other_cols = [c for c in df.columns if c not in cols]
    df = df[cols + other_cols]
    
    # Mostrar estadísticas
    encontrados = df['nombre_archivo'].notna().sum()
    no_encontrados = df['nombre_archivo'].isna().sum()
    print(f"\n✓ Archivos encontrados: {encontrados}")
    if no_encontrados > 0:
        print(f"⚠ Archivos no encontrados: {no_encontrados}")
        print("\nFilas sin mapeo:")
        sin_mapeo = df[df['nombre_archivo'].isna()][['Folder', 'subfolder']]
        print(sin_mapeo.to_string(index=False))
    
    # Guardar si se especifica output_file
    if output_file:
        print(f"\nGuardando resultado en: {output_file}")
        
        # Determinar si debe ser Excel o CSV basado en la extensión
        output_path = Path(output_file)
        if output_path.suffix.lower() in ['.xlsx', '.xls']:
            # Guardar como Excel con formato mejorado
            try:
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Datos')
                    
                    # Obtener la hoja de trabajo para formatear
                    worksheet = writer.sheets['Datos']
                    
                    # Importar función para obtener letra de columna
                    from openpyxl.utils import get_column_letter
                    
                    # Ajustar ancho de columnas automáticamente
                    for idx, col in enumerate(df.columns, start=1):
                        max_length = max(
                            df[col].astype(str).map(len).max(),
                            len(str(col))
                        ) + 2
                        # Limitar el ancho máximo a 50 caracteres
                        max_length = min(max_length, 50)
                        col_letter = get_column_letter(idx)
                        worksheet.column_dimensions[col_letter].width = max_length
                    
                    # Formatear encabezado (negrita, fondo gris)
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Congelar la primera fila (encabezado)
                    worksheet.freeze_panes = 'A2'
                    
                print("✓ Archivo Excel guardado con formato")
            except ImportError:
                print("⚠ openpyxl no está instalado. Instalando...")
                import subprocess
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                # Reintentar
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Datos')
                    worksheet = writer.sheets['Datos']
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).map(len).max(),
                            len(str(col))
                        ) + 2
                        max_length = min(max_length, 50)
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length
                print("✓ Archivo Excel guardado")
        else:
            # Guardar como CSV
            df.to_csv(output_file, index=False, sep='\t', encoding='utf-8')
            print("✓ Archivo CSV guardado")
    
    return df


def main():
    """Función principal del script."""
    if len(sys.argv) < 3:
        print("Error: Debes proporcionar el archivo de entrada y el config.yaml")
        print("\nUso:")
        print(f"  python {sys.argv[0]} <archivo_tabla> <config.yaml> [archivo_salida]")
        print("\nEjemplo:")
        print(f'  python {sys.argv[0]} tabla.txt config.yaml tabla_transformada.xlsx')
        print(f'  python {sys.argv[0]} tabla.txt config.yaml tabla_transformada.csv')
        sys.exit(1)
    
    input_file = sys.argv[1]
    config_path = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else None
    
    # Verificar que los archivos existen
    if not Path(input_file).exists():
        print(f"Error: El archivo de entrada no existe: {input_file}")
        sys.exit(1)
    
    if not Path(config_path).exists():
        print(f"Error: El archivo config.yaml no existe: {config_path}")
        sys.exit(1)
    
    try:
        df = transform_table(input_file, config_path, output_file)
        print(f"\n✓ Transformación completada!")
        print(f"  Total de filas: {len(df)}")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

